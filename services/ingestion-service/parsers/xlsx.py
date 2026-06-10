from __future__ import annotations

import io

from app.object_storage import SourceObject
from parsers.base import DocumentParser, ParsedBlock, ParserError, ParserResult, ParserUnavailable

XLSX_MIME_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel.sheet.macroEnabled.12",
}

MAX_ROWS_PER_BLOCK = 40
MAX_EMPTY_ROWS_IN_A_ROW = 50


class XlsxParser(DocumentParser):
    name = "xlsx"

    def supports(self, source: SourceObject) -> bool:
        filename = source.filename.lower()
        return source.mime_type in XLSX_MIME_TYPES or filename.endswith((".xlsx", ".xlsm"))

    def parse(self, source: SourceObject, *, parser_profile: str) -> ParserResult:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover
            raise ParserUnavailable("XLSX_PARSER_UNAVAILABLE", "openpyxl is not installed") from exc

        try:
            workbook = load_workbook(io.BytesIO(source.content), read_only=True, data_only=True)
        except Exception as exc:
            raise ParserError("XLSX_PARSE_FAILED", f"Workbook could not be opened: {exc.__class__.__name__}") from exc

        blocks: list[ParsedBlock] = []
        warnings: list[tuple[str, str]] = []
        tables_detected = 0
        offset = 0

        try:
            for sheet_index, sheet in enumerate(workbook.worksheets, start=1):
                sheet_rows: list[str] = []
                header_row: str | None = None
                empty_streak = 0
                for row in sheet.iter_rows(values_only=True):
                    cells = [_cell_text(value) for value in row]
                    if not any(cells):
                        empty_streak += 1
                        if empty_streak >= MAX_EMPTY_ROWS_IN_A_ROW:
                            break
                        continue
                    empty_streak = 0
                    line = " | ".join(cell for cell in cells if cell)
                    if header_row is None:
                        header_row = line
                    sheet_rows.append(line)

                if not sheet_rows:
                    continue
                tables_detected += 1

                for start in range(0, len(sheet_rows), MAX_ROWS_PER_BLOCK):
                    window = sheet_rows[start : start + MAX_ROWS_PER_BLOCK]
                    # Repeat the header in continuation blocks so each chunk
                    # stays interpretable on its own.
                    if start > 0 and header_row and window[0] != header_row:
                        window = [header_row, *window]
                    text = "\n".join(window)
                    blocks.append(
                        ParsedBlock(
                            text=text,
                            page_number=sheet_index,
                            section_path=[str(sheet.title)],
                            section_title=str(sheet.title),
                            article_number=None,
                            paragraph_number=None,
                            char_start=offset,
                            char_end=offset + len(text),
                            block_type="table",
                            metadata={"sheet": str(sheet.title), "row_offset": start},
                        )
                    )
                    offset += len(text) + 2
        finally:
            workbook.close()

        if not blocks:
            warnings.append(("NO_TEXT_EXTRACTED", "Workbook contains no readable cell values."))

        return ParserResult(
            parser_name=self.name,
            blocks=blocks,
            pages_processed=len(workbook.worksheets),
            tables_detected=tables_detected,
            warnings=warnings,
        )


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()
